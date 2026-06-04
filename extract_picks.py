"""
Extracts participant picks from all submission spreadsheets in the submissions/
folder and writes a Picks tab to WC2026_Pool.xlsx.

Picks tab layout:
  Row 1: Title
  Row 2: Column headers — Participant | T1 P1 | T1 P2 | ... | T6 P2 | TOTAL PTS
  Row 3+: One row per participant with their team picks and a blank Total column

Run again at any time to refresh picks from the submissions folder.
"""

import os
import glob
import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Works whether the script lives at the project root or in a scripts/ subfolder
PROJECT_FOLDER = _SCRIPT_DIR if os.path.isdir(os.path.join(_SCRIPT_DIR, "submissions")) else os.path.dirname(_SCRIPT_DIR)

SUBMISSIONS_DIR = os.path.join(PROJECT_FOLDER, "submissions")
MASTER_FILE = os.path.join(PROJECT_FOLDER, "WC2026_Pool.xlsx")
PICKS_SHEET = "Picks"

TIERS = ["Tier 1", "Tier 2", "Tier 3", "Tier 4", "Tier 5", "Tier 6"]
TIER_LABELS = {
    "Tier 1": "Elite Favorites",
    "Tier 2": "Contenders",
    "Tier 3": "Dark Horses",
    "Tier 4": "Long Shots",
    "Tier 5": "Underdogs",
    "Tier 6": "Major Underdogs",
}


def extract_picks(filepath):
    """Return (participant_name, picks_dict) from a submission file.

    picks_dict maps tier string -> [pick1, pick2]
    """
    wb = openpyxl.load_workbook(filepath)
    if "User Submission Template" not in wb.sheetnames:
        print(f"  WARNING: no 'User Submission Template' sheet in {filepath}, skipping.")
        return None, None

    ws = wb["User Submission Template"]

    # Participant name is in the row containing "Participant Name:" — col C
    participant = None
    for row in ws.iter_rows(values_only=True):
        if row[0] and str(row[0]).strip() == "Participant Name:":
            participant = row[2]
            break

    if not participant:
        print(f"  WARNING: participant name not found in {filepath}, skipping.")
        return None, None

    participant = str(participant).strip()

    # Picks: rows where col A is "Tier X" and col C is "Pick 1" or "Pick 2"
    picks = {tier: [None, None] for tier in TIERS}
    for row in ws.iter_rows(values_only=True):
        tier_val = str(row[0]).strip() if row[0] else ""
        pick_num = str(row[2]).strip() if row[2] else ""
        team = str(row[3]).strip() if row[3] else ""

        if tier_val in TIERS and pick_num in ("Pick 1", "Pick 2") and team:
            idx = 0 if pick_num == "Pick 1" else 1
            picks[tier_val][idx] = team

    return participant, picks


def build_picks_sheet():
    participants = []

    submission_files = [os.path.join(SUBMISSIONS_DIR, x) for x in os.listdir(SUBMISSIONS_DIR) if x.endswith(".xlsx")]

    if not submission_files:
        print(f"No .xlsx files found in {SUBMISSIONS_DIR}/")
        return

    for filepath in submission_files:
        print(f"Reading {os.path.basename(filepath)}...")
        name, picks = extract_picks(filepath)
        if name:
            participants.append((name, picks))

    if not participants:
        print("No valid submissions found.")
        return

    wb = openpyxl.load_workbook(MASTER_FILE)

    # Remove old or existing sheet so we can rebuild cleanly
    for old_name in ("Scoring", PICKS_SHEET):
        if old_name in wb.sheetnames:
            del wb[old_name]

    ws = wb.create_sheet(PICKS_SHEET)

    # ── Styles ──────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F4E79")   # dark blue
    tier_fills = {
        "Tier 1": PatternFill("solid", fgColor="C00000"),  # deep red
        "Tier 2": PatternFill("solid", fgColor="E26B0A"),  # orange
        "Tier 3": PatternFill("solid", fgColor="375623"),  # dark green
        "Tier 4": PatternFill("solid", fgColor="17375E"),  # navy
        "Tier 5": PatternFill("solid", fgColor="7030A0"),  # purple
        "Tier 6": PatternFill("solid", fgColor="595959"),  # grey
    }
    white_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: title ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value = "2026 FIFA World Cup Pool – Participant Picks & Scoring"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = header_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 24

    # ── Row 2: column headers ────────────────────────────────────────────────
    headers = ["Participant"]
    for tier in TIERS:
        headers.append(f"{tier}\nPick 1")
        headers.append(f"{tier}\nPick 2")
    headers.append("TOTAL PTS")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

        # Colour tier pick columns to match their tier
        tier_key = None
        for t in TIERS:
            if header.startswith(t):
                tier_key = t
                break
        cell.fill = tier_fills[tier_key] if tier_key else header_fill

    ws.row_dimensions[2].height = 30

    # ── Rows 3+: participant picks ────────────────────────────────────────────
    for row_offset, (name, picks) in enumerate(participants):
        row = 3 + row_offset
        row_bg = PatternFill("solid", fgColor="EBF3FB") if row_offset % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        name_cell = ws.cell(row=row, column=1, value=name)
        name_cell.font = bold_font
        name_cell.alignment = left
        name_cell.fill = row_bg
        name_cell.border = border

        col = 2
        for tier in TIERS:
            for pick in picks[tier]:
                cell = ws.cell(row=row, column=col, value=pick or "")
                cell.alignment = center
                cell.fill = row_bg
                cell.border = border
                col += 1

        # Total pts — blank, to be filled in manually or by a future scoring script
        total_cell = ws.cell(row=row, column=col, value=None)
        total_cell.alignment = center
        total_cell.fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow
        total_cell.border = border
        total_cell.font = Font(bold=True)

        ws.row_dimensions[row].height = 18

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 20
    for col_idx in range(2, len(headers) + 1):
        letter = get_column_letter(col_idx)
        if col_idx == len(headers):     # TOTAL PTS
            ws.column_dimensions[letter].width = 12
        else:
            ws.column_dimensions[letter].width = 16

    # Freeze the header rows and participant name column
    ws.freeze_panes = "B3"

    wb.save(MASTER_FILE)
    print(f"\nDone. Picks tab written to {MASTER_FILE} with {len(participants)} participant(s).")
    print("Participants added:")
    for name, picks in participants:
        all_picks = [p for tier in TIERS for p in picks[tier] if p]
        print(f"  {name}: {', '.join(all_picks)}")


if __name__ == "__main__":
    build_picks_sheet()
