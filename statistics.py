"""
Reads all participant picks from submissions/ and writes a Statistics tab
to WC2026_Pool.xlsx.

Sections in the Statistics tab:
  1. Team pick counts — every team, grouped by tier, sorted by pick popularity
  2. Pool Analysis — consensus lineup, participant metrics
  3. Awards & Badges — fun per-participant and per-pair awards

Run: python statistics.py
"""

import glob
import os
from collections import Counter
from itertools import combinations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Works whether the script lives at the project root or in a scripts/ subfolder
PROJECT_FOLDER = _SCRIPT_DIR if os.path.isdir(os.path.join(_SCRIPT_DIR, "submissions")) else os.path.dirname(_SCRIPT_DIR)

SUBMISSIONS_DIR = os.path.join(PROJECT_FOLDER, "submissions")
MASTER_FILE = os.path.join(PROJECT_FOLDER, "WC2026_Pool.xlsx")

STATS_SHEET = "Statistics"

TIERS = ["Tier 1", "Tier 2", "Tier 3", "Tier 4", "Tier 5", "Tier 6"]
TIER_LABELS = {
    "Tier 1": "Elite Favorites",
    "Tier 2": "Contenders",
    "Tier 3": "Dark Horses",
    "Tier 4": "Long Shots",
    "Tier 5": "Underdogs",
    "Tier 6": "Major Underdogs",
}
TIER_COLORS = {
    "Tier 1": "C00000",
    "Tier 2": "E26B0A",
    "Tier 3": "375623",
    "Tier 4": "17375E",
    "Tier 5": "7030A0",
    "Tier 6": "595959",
}


# ── Data loading ──────────────────────────────────────────────────────────────

def read_all_teams() -> list[tuple[str, str, str]]:
    """Return [(tier, team, group), ...] from the master Teams by Tier sheet."""
    wb = openpyxl.load_workbook(MASTER_FILE)
    ws = wb["Teams by Tier"]
    teams = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        tier, team, group = row[0], row[1], row[2]
        if tier and team and str(tier).startswith("Tier"):
            teams.append((str(tier).strip(), str(team).strip(), str(group).strip()))
    return teams


def read_all_picks() -> dict[str, list[str]]:
    """Return {participant_name: [team, ...]} from every submission file."""
    participants: dict[str, list[str]] = {}
    for filepath in sorted(glob.glob(os.path.join(SUBMISSIONS_DIR, "*.xlsx"))):
        wb = openpyxl.load_workbook(filepath)
        if "User Submission Template" not in wb.sheetnames:
            print(f"  WARNING: no 'User Submission Template' in {os.path.basename(filepath)}, skipping.")
            continue
        ws = wb["User Submission Template"]

        name = None
        picks: list[str] = []
        for row in ws.iter_rows(values_only=True):
            if row[0] and str(row[0]).strip() == "Participant Name:":
                name = str(row[2]).strip() if row[2] else None
            tier_val = str(row[0]).strip() if row[0] else ""
            pick_num = str(row[2]).strip() if row[2] else ""
            team = str(row[3]).strip() if row[3] else ""
            if tier_val in TIERS and pick_num in ("Pick 1", "Pick 2") and team:
                picks.append(team)

        if name:
            participants[name] = picks
        else:
            print(f"  WARNING: participant name missing in {os.path.basename(filepath)}, skipping.")

    return participants


# ── Analysis ──────────────────────────────────────────────────────────────────

def build_pick_counts(
    teams: list[tuple[str, str, str]],
    participants: dict[str, list[str]],
) -> list[dict]:
    """For every team, count pickers and who they are. Sorted tier → desc count."""
    total = len(participants)
    rows = []
    for tier, team, group in teams:
        pickers = [name for name, picks in participants.items() if team in picks]
        rows.append({
            "tier": tier,
            "team": team,
            "group": group,
            "count": len(pickers),
            "pct": len(pickers) / total * 100 if total else 0,
            "pickers": pickers,
        })
    rows.sort(key=lambda r: (TIERS.index(r["tier"]), -r["count"], r["team"]))
    return rows


def build_pick_count_map(rows: list[dict]) -> dict[str, int]:
    """Return {team: count} for quick lookup."""
    return {r["team"]: r["count"] for r in rows}


def consensus_lineup(rows: list[dict]) -> dict[str, list[str]]:
    """
    For each tier, return the teams with the highest pick count (ties included).
    Returns {tier: [team, ...]} — usually 1-2 teams per tier, could be more on ties.
    """
    consensus: dict[str, list[str]] = {}
    for tier in TIERS:
        tier_rows = [r for r in rows if r["tier"] == tier]
        if not tier_rows:
            consensus[tier] = []
            continue
        max_count = max(r["count"] for r in tier_rows)
        if max_count == 0:
            consensus[tier] = []
        else:
            consensus[tier] = [r["team"] for r in tier_rows if r["count"] == max_count]
    return consensus


def least_popular_set(rows: list[dict]) -> set[str]:
    """
    Teams with the minimum non-zero pick count across the whole pool.
    Used to identify the most contrarian picks.
    """
    counts = [r["count"] for r in rows if r["count"] > 0]
    if not counts:
        return set()
    min_count = min(counts)
    return {r["team"] for r in rows if r["count"] == min_count}


def popularity_score(picks: list[str], count_map: dict[str, int]) -> int:
    """Sum of pick counts for all of a participant's picks. Higher = more mainstream."""
    return sum(count_map.get(team, 0) for team in picks)


def uniqueness_score(picks: list[str], count_map: dict[str, int], total: int) -> float:
    """
    Average inverse-popularity across all 12 picks, scaled 0–100.
      100 = every pick was chosen by this person alone
        0 = every pick was chosen by all participants
    Formula per pick: (1 - pick_count / total_participants) × 100
    """
    if not picks or total == 0:
        return 0.0
    scores = [(1 - count_map.get(team, 0) / total) * 100 for team in picks]
    return sum(scores) / len(scores)


def consensus_overlap(picks: list[str], consensus: dict[str, list[str]]) -> int:
    """Count how many of a participant's picks appear in the consensus lineup."""
    consensus_set = {team for teams in consensus.values() for team in teams}
    return sum(1 for team in picks if team in consensus_set)


def contrarian_overlap(picks: list[str], least_pop: set[str]) -> int:
    """Count how many of a participant's picks are in the least-popular set."""
    return sum(1 for team in picks if team in least_pop)


def build_group_map(teams: list[tuple[str, str, str]]) -> dict[str, str]:
    """Return {team: wc_group} for all teams."""
    return {team: group for _, team, group in teams}


def exclusive_picks(picks: list[str], count_map: dict[str, int]) -> list[str]:
    """Teams in picks that no one else chose (count == 1)."""
    return [team for team in picks if count_map.get(team, 0) == 1]


def group_diversity(picks: list[str], group_map: dict[str, str]) -> int:
    """Count of unique WC groups represented across picks."""
    return len({group_map[t] for t in picks if t in group_map})


def group_concentration(picks: list[str], group_map: dict[str, str]) -> tuple[int, str, list[str]]:
    """
    Return (max_count, group_name, teams_from_that_group) for the WC group
    most represented in picks. Highest count = most concentrated / risky.
    """
    counts = Counter(group_map[t] for t in picks if t in group_map)
    if not counts:
        return 0, "", []
    top_group, top_count = counts.most_common(1)[0]
    teams = [t for t in picks if group_map.get(t) == top_group]
    return top_count, top_group, teams


def pairwise_similarity(participants: dict[str, list[str]]) -> list[tuple[str, str, int, float]]:
    """
    Jaccard similarity for every pair of participants.
    Returns [(name_a, name_b, shared_count, jaccard), ...] sorted descending by jaccard.
    """
    results = []
    for (a, picks_a), (b, picks_b) in combinations(participants.items(), 2):
        set_a, set_b = set(picks_a), set(picks_b)
        shared = len(set_a & set_b)
        union = len(set_a | set_b)
        results.append((a, b, shared, shared / union if union else 0.0))
    results.sort(key=lambda x: -x[3])
    return results


def compute_badges(
    participants: dict[str, list[str]],
    rows: list[dict],
    teams: list[tuple[str, str, str]],
) -> list[dict]:
    """
    Compute all badge winners. Returns a list of badge dicts ready for rendering.
    Each dict: name, icon, description, winners, stat, bg (hex), fg (hex).
    """
    count_map = build_pick_count_map(rows)
    group_map = build_group_map(teams)
    total = len(participants)

    metrics = {
        name: {
            "popularity":    popularity_score(picks, count_map),
            "uniqueness":    uniqueness_score(picks, count_map, total),
            "exclusive":     exclusive_picks(picks, count_map),
            "diversity":     group_diversity(picks, group_map),
            "concentration": group_concentration(picks, group_map),
        }
        for name, picks in participants.items()
    }

    def winners_of(key, best=max):
        target = best(m[key] for m in metrics.values())
        return [n for n, m in metrics.items() if m[key] == target], target

    badges = []

    # ── Crowd Favourite ───────────────────────────────────────────────────────
    w, val = winners_of("popularity", best=max)
    badges.append({
        "icon": "⭐", "name": "Crowd Favourite",
        "description": "Most mainstream picks — highest overlap with the popular consensus",
        "winners": w, "stat": f"Popularity score: {val}",
        "bg": "FFF2CC", "fg": "7F6000",
    })

    # ── Dark Horse ───────────────────────────────────────────────────────────
    max_u = max(m["uniqueness"] for m in metrics.values())
    dark_w = [n for n, m in metrics.items() if abs(m["uniqueness"] - max_u) < 0.01]
    badges.append({
        "icon": "🎯", "name": "Dark Horse",
        "description": "Most contrarian picks — highest uniqueness score",
        "winners": dark_w, "stat": f"Uniqueness: {max_u:.1f}%",
        "bg": "EAD1DC", "fg": "4A235A",
    })

    # ── Lone Wolf ────────────────────────────────────────────────────────────
    max_excl = max((len(m["exclusive"]) for m in metrics.values()), default=0)
    wolf_w = [n for n, m in metrics.items() if len(m["exclusive"]) == max_excl]
    wolf_teams = {t for n in wolf_w for t in metrics[n]["exclusive"]}
    badges.append({
        "icon": "🐺", "name": "Lone Wolf",
        "description": "Most picks that nobody else made",
        "winners": wolf_w if max_excl > 0 else [],
        "stat": (f"{max_excl} exclusive pick(s): {', '.join(sorted(wolf_teams))}"
                 if max_excl > 0 else "No exclusive picks in the pool"),
        "bg": "D0E0E3", "fg": "0C343D",
    })

    # ── Globetrotter ─────────────────────────────────────────────────────────
    max_div = max(m["diversity"] for m in metrics.values())
    globe_w = [n for n, m in metrics.items() if m["diversity"] == max_div]
    badges.append({
        "icon": "🌍", "name": "Globetrotter",
        "description": "Picks span the most different WC groups",
        "winners": globe_w, "stat": f"{max_div} different WC groups covered",
        "bg": "D9EAD3", "fg": "274E13",
    })

    # ── Group Gambler ─────────────────────────────────────────────────────────
    max_conc = max(m["concentration"][0] for m in metrics.values())
    gambler_w = [n for n, m in metrics.items() if m["concentration"][0] == max_conc]
    gambler_groups = sorted({metrics[n]["concentration"][1] for n in gambler_w})
    gambler_teams = sorted({t for n in gambler_w for t in metrics[n]["concentration"][2]})
    badges.append({
        "icon": "🎲", "name": "Group Gambler",
        "description": "Most picks from the same WC group — highest concentration risk",
        "winners": gambler_w,
        "stat": f"{max_conc} picks from {' / '.join(gambler_groups)}: {', '.join(gambler_teams)}",
        "bg": "FCE5CD", "fg": "7F2B00",
    })

    # ── Twins ────────────────────────────────────────────────────────────────
    pairs = pairwise_similarity(participants)
    if pairs:
        a, b, shared, jaccard = pairs[0]
        shared_teams = sorted(set(participants[a]) & set(participants[b]))
        badges.append({
            "icon": "🤝", "name": "Twins",
            "description": "The pair of participants with the most picks in common",
            "winners": [f"{a} & {b}"],
            "stat": f"{shared} shared picks ({jaccard:.0%} similarity): {', '.join(shared_teams)}",
            "bg": "CFE2F3", "fg": "1C4587",
        })

    return badges


# ── Spreadsheet helpers ───────────────────────────────────────────────────────

def _lighten(hex_color: str, factor: float = 0.82) -> str:
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"{r:02X}{g:02X}{b:02X}"


def _section_header(ws, row: int, text: str, n_cols: int, fill_color: str = "1F4E79"):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    return row + 1


def _col_header_row(ws, row: int, headers: list[str], fill_color: str = "1F4E79"):
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[row].height = 18
    return row + 1


def _blank_row(ws, row: int, n_cols: int):
    ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
    ws.row_dimensions[row].height = 8
    return row + 1


def write_badges_section(ws, data_row: int, badges: list[dict], N_COLS: int) -> int:
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    data_row = _blank_row(ws, data_row, N_COLS)
    data_row = _section_header(ws, data_row, "  AWARDS & BADGES", N_COLS, fill_color="BF9000")

    # Column headers
    data_row = _col_header_row(
        ws, data_row,
        ["Badge", "What it means", "", "Winner(s)", "", "Earning stat"],
        fill_color="BF9000",
    )

    for badge in badges:
        fill = PatternFill("solid", fgColor=badge["bg"])
        name_font = Font(bold=True, color=badge["fg"])
        stat_font = Font(italic=True, color=badge["fg"])
        winner_font = Font(bold=True, color=badge["fg"])
        desc_font = Font(italic=True, color="595959", size=9)

        # Col A: icon + badge name
        ws.cell(row=data_row, column=1,
                value=f"{badge['icon']}  {badge['name']}").font = name_font
        ws.cell(row=data_row, column=1).fill = fill
        ws.cell(row=data_row, column=1).alignment = left
        ws.cell(row=data_row, column=1).border = border

        # Col B-C: description
        ws.merge_cells(f"B{data_row}:C{data_row}")
        cell = ws.cell(row=data_row, column=2, value=badge["description"])
        cell.font = desc_font
        cell.fill = fill
        cell.alignment = left
        cell.border = border

        # Col D-E: winner(s)
        ws.merge_cells(f"D{data_row}:E{data_row}")
        winners_str = "  /  ".join(badge["winners"]) if badge["winners"] else "—"
        cell = ws.cell(row=data_row, column=4, value=winners_str)
        cell.font = winner_font
        cell.fill = fill
        cell.alignment = left
        cell.border = border

        # Col F: stat
        cell = ws.cell(row=data_row, column=6, value=badge["stat"])
        cell.font = stat_font
        cell.fill = fill
        cell.alignment = left
        cell.border = border

        ws.row_dimensions[data_row].height = 22
        data_row += 1

    return data_row


def write_statistics_sheet(
    rows: list[dict],
    participants: dict[str, list[str]],
    teams: list[tuple[str, str, str]],
):
    total = len(participants)
    wb = openpyxl.load_workbook(MASTER_FILE)

    if STATS_SHEET in wb.sheetnames:
        del wb[STATS_SHEET]
    ws = wb.create_sheet(STATS_SHEET)

    N_COLS = 6
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    alt_fills = {
        tier: PatternFill("solid", fgColor=_lighten(TIER_COLORS[tier]))
        for tier in TIERS
    }
    tier_header_fills = {
        tier: PatternFill("solid", fgColor=TIER_COLORS[tier])
        for tier in TIERS
    }

    # ── Section 1: title + team pick counts ──────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    t = ws["A1"]
    t.value = f"2026 FIFA World Cup Pool – Team Pick Statistics  ({total} participants)"
    t.font = Font(bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="1F4E79")
    t.alignment = center
    ws.row_dimensions[1].height = 22

    data_row = _col_header_row(ws, 2, ["Tier", "Team", "WC Group", "# Picked", "% Picked", "Picked By"])

    current_tier = None
    for entry in rows:
        tier = entry["tier"]
        if tier != current_tier:
            current_tier = tier
            ws.merge_cells(f"A{data_row}:{get_column_letter(N_COLS)}{data_row}")
            hdr = ws.cell(row=data_row, column=1,
                          value=f"  {tier} – {TIER_LABELS[tier]}")
            hdr.font = Font(bold=True, color="FFFFFF", size=11)
            hdr.fill = tier_header_fills[tier]
            hdr.alignment = left
            hdr.border = border
            ws.row_dimensions[data_row].height = 18
            data_row += 1

        row_fill = alt_fills[tier]
        values = [
            tier,
            entry["team"],
            entry["group"],
            entry["count"],
            f"{entry['pct']:.0f}%",
            ", ".join(entry["pickers"]) if entry["pickers"] else "—",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = center if col != 6 else left
        if entry["count"] > 0:
            ws.cell(row=data_row, column=4).font = Font(bold=True)
        ws.row_dimensions[data_row].height = 16
        data_row += 1

    # ── Section 2: Pool Analysis ──────────────────────────────────────────────
    data_row = _blank_row(ws, data_row, N_COLS)
    data_row = _section_header(ws, data_row, "  POOL ANALYSIS", N_COLS)
    data_row = _blank_row(ws, data_row, N_COLS)

    count_map = build_pick_count_map(rows)
    consensus = consensus_lineup(rows)
    least_pop = least_popular_set(rows)

    # ── 2a. Consensus lineup ─────────────────────────────────────────────────
    data_row = _section_header(
        ws, data_row,
        "  1.  Consensus Lineup  –  most popular pick(s) per tier",
        N_COLS, fill_color="2E4057",
    )
    data_row = _col_header_row(
        ws, data_row,
        ["Tier", "Tier Description", "Consensus Pick(s)", "Times Picked", "% of Pool", ""],
        fill_color="4472C4",
    )
    for tier in TIERS:
        con_teams = consensus.get(tier, [])
        team_str = "  /  ".join(con_teams) if con_teams else "—  (no picks)"
        count_val = count_map.get(con_teams[0], 0) if con_teams else 0
        pct_val = f"{count_val / total * 100:.0f}%" if total and con_teams else "—"
        fill = PatternFill("solid", fgColor="DCE6F1")
        vals = [tier, TIER_LABELS[tier], team_str, count_val or "", pct_val, ""]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = center if col not in (2, 3) else left
        if con_teams:
            ws.cell(row=data_row, column=3).font = Font(bold=True)
        ws.row_dimensions[data_row].height = 16
        data_row += 1

    data_row = _blank_row(ws, data_row, N_COLS)

    # ── 2b & 2c & 2d. Participant analysis table ──────────────────────────────
    data_row = _section_header(
        ws, data_row,
        "  2 – 4.  Participant Analysis",
        N_COLS, fill_color="2E4057",
    )

    # Compute metrics for all participants
    analysis = []
    for name, picks in participants.items():
        pop_score = popularity_score(picks, count_map)
        u_score = uniqueness_score(picks, count_map, total)
        c_overlap = consensus_overlap(picks, consensus)
        contra_overlap = contrarian_overlap(picks, least_pop)
        analysis.append({
            "name": name,
            "picks": picks,
            "popularity_score": pop_score,
            "uniqueness_score": u_score,
            "consensus_overlap": c_overlap,
            "contrarian_overlap": contra_overlap,
        })

    max_pop = max(a["popularity_score"] for a in analysis)
    min_pop = min(a["popularity_score"] for a in analysis)
    max_unique = max(a["uniqueness_score"] for a in analysis)

    most_mainstream = [a for a in analysis if a["popularity_score"] == max_pop]
    most_contrarian = [a for a in analysis if a["popularity_score"] == min_pop]

    # Explanation note row
    notes = [
        ("Popularity Score",
         f"Sum of how many participants picked each of your 12 teams (max = {total * 12}).  Higher = more mainstream."),
        ("Uniqueness Score",
         f"Average rarity of your picks, 0–100.  Per pick: (1 − pick_count ÷ {total}) × 100.  Higher = more unique."),
        ("Most Mainstream",
         "Participant(s) with the highest popularity score — their picks most overlap the popular consensus."),
        ("Most Contrarian",
         "Participant(s) with the lowest popularity score — their picks most overlap the least-chosen teams."),
    ]
    note_fill = PatternFill("solid", fgColor="F2F2F2")
    for label, note in notes:
        ws.cell(row=data_row, column=1, value=label).font = Font(bold=True, italic=True, size=9)
        ws.cell(row=data_row, column=1).fill = note_fill
        ws.cell(row=data_row, column=1).alignment = left
        ws.merge_cells(f"B{data_row}:{get_column_letter(N_COLS)}{data_row}")
        note_cell = ws.cell(row=data_row, column=2, value=note)
        note_cell.font = Font(italic=True, size=9, color="595959")
        note_cell.fill = note_fill
        note_cell.alignment = left
        ws.row_dimensions[data_row].height = 14
        data_row += 1

    data_row = _blank_row(ws, data_row, N_COLS)

    # Participant analysis table headers
    data_row = _col_header_row(
        ws, data_row,
        ["Participant", "Popularity\nScore", "Uniqueness\nScore", "Consensus\nOverlap", "Contrarian\nOverlap", "Profile"],
        fill_color="4472C4",
    )
    ws.row_dimensions[data_row - 1].height = 28

    analysis_sorted = sorted(analysis, key=lambda a: -a["popularity_score"])
    for i, a in enumerate(analysis_sorted):
        fill = PatternFill("solid", fgColor="DCE6F1" if i % 2 == 0 else "FFFFFF")

        is_mainstream = a["popularity_score"] == max_pop
        is_contrarian = a["popularity_score"] == min_pop
        is_most_unique = abs(a["uniqueness_score"] - max_unique) < 0.01

        labels = []
        if is_mainstream:
            labels.append("Most Mainstream")
        if is_contrarian:
            labels.append("Most Contrarian")
        if is_most_unique and not (is_mainstream or is_contrarian):
            labels.append("Most Unique")
        profile = "  |  ".join(labels) if labels else "—"

        vals = [
            a["name"],
            a["popularity_score"],
            f"{a['uniqueness_score']:.1f}",
            a["consensus_overlap"],
            a["contrarian_overlap"],
            profile,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = left if col in (1, 6) else center

        # Bold name, colour profile badges
        ws.cell(row=data_row, column=1).font = Font(bold=True)
        profile_cell = ws.cell(row=data_row, column=6)
        if is_mainstream:
            profile_cell.font = Font(bold=True, color="1F4E79")
        elif is_contrarian:
            profile_cell.font = Font(bold=True, color="7030A0")

        ws.row_dimensions[data_row].height = 16
        data_row += 1

    data_row = _blank_row(ws, data_row, N_COLS)

    # ── 2b/2c highlight callout rows ─────────────────────────────────────────
    callouts = [
        (
            "Most like the popular consensus:",
            "  /  ".join(a["name"] for a in most_mainstream),
            "DCE6F1",
            "1F4E79",
        ),
        (
            "Most contrarian picks:",
            "  /  ".join(a["name"] for a in most_contrarian),
            "E8D5F5",
            "7030A0",
        ),
    ]
    for label, value, bg, fg in callouts:
        ws.merge_cells(f"A{data_row}:C{data_row}")
        lbl = ws.cell(row=data_row, column=1, value=label)
        lbl.font = Font(bold=True, size=10, color=fg)
        lbl.fill = PatternFill("solid", fgColor=bg)
        lbl.alignment = left
        ws.merge_cells(f"D{data_row}:{get_column_letter(N_COLS)}{data_row}")
        val = ws.cell(row=data_row, column=4, value=value)
        val.font = Font(bold=True, size=10, color=fg)
        val.fill = PatternFill("solid", fgColor=bg)
        val.alignment = left
        ws.row_dimensions[data_row].height = 18
        data_row += 1

    # ── Section 3: Badges ─────────────────────────────────────────────────────
    badges = compute_badges(participants, rows, teams)
    data_row = write_badges_section(ws, data_row, badges, N_COLS)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 5
    ws.column_dimensions["F"].width = 52

    ws.freeze_panes = "A3"
    wb.save(MASTER_FILE)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Reading teams from master file...")
    teams = read_all_teams()
    print(f"  {len(teams)} teams found.")

    print("Reading participant picks from submissions/...")
    participants = read_all_picks()
    print(f"  {len(participants)} participant(s): {', '.join(participants)}")

    print("Building pick counts...")
    rows = build_pick_counts(teams, participants)
    count_map = build_pick_count_map(rows)

    picked = [r for r in rows if r["count"] > 0]
    unpicked = [r for r in rows if r["count"] == 0]
    print(f"  {len(picked)} team(s) picked, {len(unpicked)} not picked.")

    print("Writing Statistics sheet...")
    write_statistics_sheet(rows, participants, teams)
    print(f"Done. '{STATS_SHEET}' tab written to {MASTER_FILE}.")

    # ── Console summary ───────────────────────────────────────────────────────
    total = len(participants)
    consensus = consensus_lineup(rows)
    least_pop = least_popular_set(rows)

    print("\n── Consensus lineup (most picked per tier) ──")
    for tier in TIERS:
        teams_str = " / ".join(consensus.get(tier, ["—"]))
        print(f"  {tier}: {teams_str}")

    print("\n── Participant analysis ──")
    for name, picks in participants.items():
        pop = popularity_score(picks, count_map)
        u = uniqueness_score(picks, count_map, total)
        c_ov = consensus_overlap(picks, consensus)
        contra_ov = contrarian_overlap(picks, least_pop)
        print(f"  {name:<20}  popularity={pop:2d}  uniqueness={u:5.1f}%  "
              f"consensus_overlap={c_ov}  contrarian_overlap={contra_ov}")

    analysis = [
        {"name": n, "popularity_score": popularity_score(p, count_map)}
        for n, p in participants.items()
    ]
    mainstream = [a["name"] for a in analysis if a["popularity_score"] == max(a["popularity_score"] for a in analysis)]
    contrarian = [a["name"] for a in analysis if a["popularity_score"] == min(a["popularity_score"] for a in analysis)]
    print(f"\n  Most mainstream : {' / '.join(mainstream)}")
    print(f"  Most contrarian : {' / '.join(contrarian)}")

    print("\n── Badges ──")
    badges = compute_badges(participants, rows, teams)
    for b in badges:
        w = " / ".join(b["winners"]) if b["winners"] else "—"
        print(f"  {b['icon']} {b['name']:<20}  {w:<25}  {b['stat']}")

    print("\nPick counts by tier:")
    current_tier = None
    for r in rows:
        if r["tier"] != current_tier:
            current_tier = r["tier"]
            print(f"\n  {r['tier']} – {TIER_LABELS[r['tier']]}:")
        bar = "█" * r["count"]
        pickers = f"  ← {', '.join(r['pickers'])}" if r["pickers"] else ""
        print(f"    {r['team']:<22} {bar:<6} {r['count']}/{total}{pickers}")


if __name__ == "__main__":
    main()
