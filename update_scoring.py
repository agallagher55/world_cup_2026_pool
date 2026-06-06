"""
Creates/refreshes two tabs in WC2026_Pool.xlsx:

  Scores  — manual data-entry sheet pre-populated with all 72 group stage
             matchups and 32 knockout placeholder rows.
             Columns: Match# | Stage | Group | Date | Home | Home Score |
                      Away Score | Away | Duration | Notes

  Scoring — computed leaderboard + per-pick points breakdown.
             Reads finished rows from Scores tab, applies pool scoring rules,
             and cross-references picks from the Picks tab.

Usage:
  python update_scoring.py          # rebuild both tabs
  python update_scoring.py --scores # rebuild Scores tab only
  python update_scoring.py --scoring # rebuild Scoring tab only
"""

import os
import sys
import argparse
from itertools import combinations

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Install openpyxl:  pip install openpyxl")

# ── Paths ────────────────────────────────────────────────────────────────────
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_FOLDER = (
    _SCRIPT_DIR
    if os.path.isfile(os.path.join(_SCRIPT_DIR, "WC2026_Pool.xlsx"))
    else os.path.dirname(_SCRIPT_DIR)
)
MASTER_FILE = os.path.join(PROJECT_FOLDER, "WC2026_Pool.xlsx")

SCORES_SHEET = "Scores"
SCORING_SHEET = "Scoring"
PICKS_SHEET = "Picks"

TIERS = ["Tier 1", "Tier 2", "Tier 3", "Tier 4", "Tier 5", "Tier 6"]

# ── Group compositions ───────────────────────────────────────────────────────
GROUPS = {
    "A": ["Mexico", "South Korea", "South Africa", "Czechia"],
    "B": ["Canada", "Switzerland", "Bosnia & Herz.", "Qatar"],
    "C": ["Brazil", "Morocco", "Scotland", "Haiti"],
    "D": ["United States", "Turkey", "Australia", "Paraguay"],
    "E": ["Germany", "Ecuador", "Ivory Coast", "Curaçao"],
    "F": ["Netherlands", "Japan", "Sweden", "Tunisia"],
    "G": ["Belgium", "Egypt", "Iran", "New Zealand"],
    "H": ["Spain", "Uruguay", "Saudi Arabia", "Cape Verde"],
    "I": ["France", "Senegal", "Norway", "Iraq"],
    "J": ["Argentina", "Austria", "Algeria", "Jordan"],
    "K": ["Portugal", "Colombia", "Uzbekistan", "DR Congo"],
    "L": ["England", "Croatia", "Ghana", "Panama"],
}

KNOCKOUT_ROUNDS = [
    ("Round of 32", 16),
    ("Round of 16", 8),
    ("Quarter-final", 4),
    ("Semi-final", 2),
    ("Third Place", 1),
    ("Final", 1),
]

# ── Scoring constants ─────────────────────────────────────────────────────────
PTS_WIN = 3
PTS_DRAW = 1
PTS_LOSS = 0
PTS_WIN_AET = 2   # win via extra time/penalties
PTS_LOSS_AET = 1  # loss via extra time/penalties
PTS_GROUP_WIN = 3
PTS_MOST_GOALS = 3
PTS_FEWEST_CONCEDED = 3
PTS_BEST_ATTACK_TOURNEY = 3
PTS_BEST_DEFENSE_TOURNEY = 3

# ── Style helpers ─────────────────────────────────────────────────────────────
def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
YELLOW_FILL  = PatternFill("solid", fgColor="FFF2CC")
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
GREEN_FILL   = PatternFill("solid", fgColor="E2EFDA")
GREY_FILL    = PatternFill("solid", fgColor="D9D9D9")

GROUP_FILLS = {
    "A": PatternFill("solid", fgColor="FCE4D6"),
    "B": PatternFill("solid", fgColor="FFF2CC"),
    "C": PatternFill("solid", fgColor="E2EFDA"),
    "D": PatternFill("solid", fgColor="DDEBF7"),
    "E": PatternFill("solid", fgColor="EAD1DC"),
    "F": PatternFill("solid", fgColor="D9EAD3"),
    "G": PatternFill("solid", fgColor="CFE2F3"),
    "H": PatternFill("solid", fgColor="FDE9D9"),
    "I": PatternFill("solid", fgColor="F4CCCC"),
    "J": PatternFill("solid", fgColor="D9D9D9"),
    "K": PatternFill("solid", fgColor="EAD1DC"),
    "L": PatternFill("solid", fgColor="D0E0E3"),
}

WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD       = Font(bold=True)
CENTER     = Alignment(horizontal="center", vertical="center")
LEFT       = Alignment(horizontal="left", vertical="center")


def _apply(cell, *, value=None, font=None, fill=None, align=CENTER, border=True):
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align
    if border:
        cell.border = _thin_border()


# ── Build Scores tab ──────────────────────────────────────────────────────────

def build_scores_tab(ws):
    """Write the Scores tab skeleton — group games + knockout placeholders."""

    COL_HEADERS = [
        "Match #", "Stage", "Group", "Date",
        "Home Team", "Home Score", "Away Score", "Away Team",
        "Duration", "Notes",
    ]
    COL_WIDTHS = [9, 16, 8, 16, 20, 12, 12, 20, 16, 24]

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(COL_HEADERS))}1")
    title = ws["A1"]
    title.value = "2026 FIFA World Cup – Match Scores"
    title.font  = Font(bold=True, size=14, color="FFFFFF")
    title.fill  = HEADER_FILL
    title.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # Header row
    for ci, h in enumerate(COL_HEADERS, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font  = WHITE_BOLD
        c.fill  = HEADER_FILL
        c.alignment = CENTER
        c.border = _thin_border()
    ws.row_dimensions[2].height = 22

    # Column widths
    for ci, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    row = 3
    match_num = 1

    # Group stage — section header + 6 games per group
    for grp_letter, teams in GROUPS.items():
        # Group section header
        ws.merge_cells(f"A{row}:{get_column_letter(len(COL_HEADERS))}{row}")
        hdr = ws.cell(row=row, column=1,
                      value=f"Group {grp_letter}  –  {' | '.join(teams)}")
        hdr.font  = Font(bold=True, color="FFFFFF")
        hdr.fill  = GROUP_FILLS.get(grp_letter, HEADER_FILL)
        hdr.alignment = LEFT
        ws.row_dimensions[row].height = 18
        row += 1

        fill = GROUP_FILLS.get(grp_letter, WHITE_FILL)
        for home, away in combinations(teams, 2):
            vals = [match_num, "Group Stage", f"Group {grp_letter}", "",
                    home, None, None, away, "REGULAR", ""]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                # Score cells get yellow fill for visibility
                if ci in (6, 7):
                    c.fill   = YELLOW_FILL
                    c.font   = BOLD
                elif ci == 9:  # Duration
                    c.fill   = WHITE_FILL
                else:
                    c.fill   = fill
                c.alignment = CENTER
                c.border = _thin_border()
            ws.row_dimensions[row].height = 16
            match_num += 1
            row += 1

        row += 1  # blank spacer between groups

    # Knockout rounds — section headers + placeholder rows
    ws.merge_cells(f"A{row}:{get_column_letter(len(COL_HEADERS))}{row}")
    ko_hdr = ws.cell(row=row, column=1, value="KNOCKOUT ROUNDS")
    ko_hdr.font  = Font(bold=True, size=12, color="FFFFFF")
    ko_hdr.fill  = PatternFill("solid", fgColor="243F60")
    ko_hdr.alignment = CENTER
    ws.row_dimensions[row].height = 22
    row += 1

    for round_name, n_games in KNOCKOUT_ROUNDS:
        # Round header
        ws.merge_cells(f"A{row}:{get_column_letter(len(COL_HEADERS))}{row}")
        r_hdr = ws.cell(row=row, column=1, value=round_name)
        r_hdr.font  = Font(bold=True, color="FFFFFF")
        r_hdr.fill  = PatternFill("solid", fgColor="375623")
        r_hdr.alignment = LEFT
        ws.row_dimensions[row].height = 18
        row += 1

        for _ in range(n_games):
            vals = [match_num, round_name, "", "", "", None, None, "", "REGULAR", ""]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                if ci in (6, 7):
                    c.fill = YELLOW_FILL
                    c.font = BOLD
                else:
                    c.fill = ALT_FILL if match_num % 2 == 0 else WHITE_FILL
                c.alignment = CENTER
                c.border = _thin_border()
            ws.row_dimensions[row].height = 16
            match_num += 1
            row += 1

        row += 1  # spacer

    ws.freeze_panes = "A3"
    print(f"  Scores tab: {match_num - 1} rows written.")


# ── Read finished scores ──────────────────────────────────────────────────────

def read_scores(wb):
    """Return list of finished match dicts from the Scores tab."""
    if SCORES_SHEET not in wb.sheetnames:
        return []

    ws = wb[SCORES_SHEET]
    matches = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            int(row[0])
        except (TypeError, ValueError):
            continue  # section header row

        stage      = str(row[1]).strip() if row[1] else ""
        group      = str(row[2]).strip() if row[2] else ""
        home       = str(row[4]).strip() if row[4] else ""
        home_score = row[5]
        away_score = row[6]
        away       = str(row[7]).strip() if row[7] else ""
        duration   = str(row[8]).strip().upper() if row[8] else "REGULAR"

        # Only include fully scored rows
        if home_score is None or away_score is None or not home or not away:
            continue
        try:
            hs = int(home_score)
            aws = int(away_score)
        except (TypeError, ValueError):
            continue

        matches.append({
            "stage": stage,
            "group": group,
            "home": home,
            "away": away,
            "home_score": hs,
            "away_score": aws,
            "duration": duration,
        })

    return matches


# ── Compute team stats from finished matches ──────────────────────────────────

def compute_team_stats(matches):
    """Return {team: {gf, ga, wins, draws, losses, group, stage_reached}}."""
    stats = {}

    def _ensure(team, group=""):
        if team not in stats:
            stats[team] = {
                "gf": 0, "ga": 0, "wins": 0, "draws": 0, "losses": 0,
                "group": group, "group_winner": False,
            }

    for m in matches:
        hs, aws = m["home_score"], m["away_score"]
        home, away = m["home"], m["away"]
        grp = m["group"]
        _ensure(home, grp)
        _ensure(away, grp)

        stats[home]["gf"] += hs
        stats[home]["ga"] += aws
        stats[away]["gf"] += aws
        stats[away]["ga"] += hs

        if m["stage"] == "Group Stage":
            if hs > aws:
                stats[home]["wins"] += 1
                stats[away]["losses"] += 1
            elif hs < aws:
                stats[away]["wins"] += 1
                stats[home]["losses"] += 1
            else:
                stats[home]["draws"] += 1
                stats[away]["draws"] += 1

    return stats


def apply_group_bonuses(team_stats, matches):
    """Mark group winners + best attack/defense; return bonus dict."""
    bonuses = {t: 0 for t in team_stats}

    # Determine which groups are fully played (6 games each)
    group_games = {}
    for m in matches:
        if m["stage"] == "Group Stage":
            g = m["group"]
            group_games[g] = group_games.get(g, 0) + 1

    complete_groups = {g for g, cnt in group_games.items() if cnt == 6}

    # Per-group bonuses
    tournament_best_gf = -1
    tournament_best_ga = float("inf")
    tournament_best_gf_teams = []
    tournament_best_ga_teams = []

    for grp in complete_groups:
        grp_teams = [t for t, s in team_stats.items() if s["group"] == grp]
        if not grp_teams:
            continue

        # Group winner: most wins; tie-break: gf-ga; give all tied teams bonus
        max_wins = max(team_stats[t]["wins"] for t in grp_teams)
        leaders = [t for t in grp_teams if team_stats[t]["wins"] == max_wins]
        for t in leaders:
            bonuses[t] += PTS_GROUP_WIN
            team_stats[t]["group_winner"] = True

        # Most goals scored
        max_gf = max(team_stats[t]["gf"] for t in grp_teams)
        for t in grp_teams:
            if team_stats[t]["gf"] == max_gf:
                bonuses[t] += PTS_MOST_GOALS
        if max_gf > tournament_best_gf:
            tournament_best_gf = max_gf
            tournament_best_gf_teams = [t for t in grp_teams if team_stats[t]["gf"] == max_gf]
        elif max_gf == tournament_best_gf:
            tournament_best_gf_teams += [t for t in grp_teams if team_stats[t]["gf"] == max_gf]

        # Fewest goals conceded
        min_ga = min(team_stats[t]["ga"] for t in grp_teams)
        for t in grp_teams:
            if team_stats[t]["ga"] == min_ga:
                bonuses[t] += PTS_FEWEST_CONCEDED
        if min_ga < tournament_best_ga:
            tournament_best_ga = min_ga
            tournament_best_ga_teams = [t for t in grp_teams if team_stats[t]["ga"] == min_ga]
        elif min_ga == tournament_best_ga:
            tournament_best_ga_teams += [t for t in grp_teams if team_stats[t]["ga"] == min_ga]

    # Tournament-wide attack/defense bonuses
    for t in tournament_best_gf_teams:
        bonuses[t] += PTS_BEST_ATTACK_TOURNEY
    for t in tournament_best_ga_teams:
        bonuses[t] += PTS_BEST_DEFENSE_TOURNEY

    return bonuses


def compute_match_points(team, matches, team_stats):
    """Return points earned from individual match results for a team."""
    pts = 0
    for m in matches:
        if team not in (m["home"], m["away"]):
            continue
        is_home = m["home"] == team
        my_score = m["home_score"] if is_home else m["away_score"]
        op_score = m["away_score"] if is_home else m["home_score"]
        dur = m["duration"]

        if m["stage"] == "Group Stage":
            if my_score > op_score:
                pts += PTS_WIN
            elif my_score == op_score:
                pts += PTS_DRAW
            # loss = 0
        else:
            # Knockout — no draws in final result
            if my_score > op_score:
                pts += PTS_WIN_AET if dur in ("EXTRA_TIME", "PENALTY_SHOOTOUT") else PTS_WIN
            else:
                pts += PTS_LOSS_AET if dur in ("EXTRA_TIME", "PENALTY_SHOOTOUT") else PTS_LOSS
    return pts


# ── Read participant picks from Picks tab ─────────────────────────────────────

def read_picks(wb):
    """Return list of (name, [team, ...]) from the Picks tab."""
    if PICKS_SHEET not in wb.sheetnames:
        print(f"WARNING: '{PICKS_SHEET}' tab not found. Run extract_picks.py first.")
        return []

    ws = wb[PICKS_SHEET]
    participants = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        picks = [str(c).strip() for c in row[1:] if c and str(c).strip() not in ("", "None")]
        # Drop the TOTAL PTS cell (last column may be numeric or None)
        if picks and picks[-1].replace(".", "").isdigit():
            picks = picks[:-1]
        if name and picks:
            participants.append((name, picks))

    return participants


# ── Build Scoring tab ─────────────────────────────────────────────────────────

def build_scoring_tab(ws, wb):
    matches = read_scores(wb)
    participants = read_picks(wb)

    if not participants:
        ws["A1"].value = "No participant picks found. Run extract_picks.py first."
        return

    # Compute stats
    team_stats = compute_team_stats(matches)
    bonuses = apply_group_bonuses(team_stats, matches) if matches else {}

    # Score each participant
    scored = []
    for name, picks in participants:
        total = 0
        pick_pts = {}
        for team in picks:
            mp = compute_match_points(team, matches, team_stats)
            bp = bonuses.get(team, 0)
            pick_pts[team] = mp + bp
            total += pick_pts[team]
        scored.append((name, picks, pick_pts, total))

    # Sort by total descending
    scored.sort(key=lambda x: x[3], reverse=True)

    # ── Styles ───────────────────────────────────────────────────────────────
    ALL_TEAMS = [t for _, picks, _, _ in scored for t in picks]
    # unique sorted list of picks across all participants
    unique_teams = sorted(set(ALL_TEAMS))

    tier_fills = {
        "Tier 1": PatternFill("solid", fgColor="C00000"),
        "Tier 2": PatternFill("solid", fgColor="E26B0A"),
        "Tier 3": PatternFill("solid", fgColor="375623"),
        "Tier 4": PatternFill("solid", fgColor="17375E"),
        "Tier 5": PatternFill("solid", fgColor="7030A0"),
        "Tier 6": PatternFill("solid", fgColor="595959"),
    }

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    tc = ws["A1"]
    tc.value = "2026 FIFA World Cup Pool – Scoring"
    tc.font  = Font(bold=True, size=14, color="FFFFFF")
    tc.fill  = HEADER_FILL
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 24

    # ── Leaderboard ────────────────────────────────────────────────────────────
    row = 2
    ws.merge_cells(f"A{row}:F{row}")
    lh = ws.cell(row=row, column=1, value="LEADERBOARD")
    lh.font  = Font(bold=True, size=12, color="FFFFFF")
    lh.fill  = PatternFill("solid", fgColor="243F60")
    lh.alignment = CENTER
    ws.row_dimensions[row].height = 20
    row += 1

    for ci, h in enumerate(["Rank", "Participant", "Total Points", "Pts from Matches",
                             "Pts from Bonuses", "# Scored Teams"], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font  = WHITE_BOLD
        c.fill  = HEADER_FILL
        c.alignment = CENTER
        c.border = _thin_border()
    ws.row_dimensions[row].height = 20
    row += 1

    for rank, (name, picks, pick_pts, total) in enumerate(scored, 1):
        match_pts = sum(compute_match_points(t, matches, team_stats) for t in picks)
        bonus_pts = total - match_pts
        n_scored  = sum(1 for t in picks if pick_pts.get(t, 0) > 0 or t in team_stats)
        fill = ALT_FILL if rank % 2 == 0 else WHITE_FILL
        if rank == 1:
            fill = PatternFill("solid", fgColor="FFD700")  # gold for leader
        row_vals = [rank, name, total, match_pts, bonus_pts, n_scored]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.fill = fill
            c.alignment = CENTER if ci != 2 else LEFT
            c.border = _thin_border()
            if ci in (1, 3):
                c.font = BOLD
        row += 1

    row += 1  # spacer

    # ── Per-participant breakdown ──────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    bh = ws.cell(row=row, column=1, value="PICKS BREAKDOWN")
    bh.font  = Font(bold=True, size=12, color="FFFFFF")
    bh.fill  = PatternFill("solid", fgColor="243F60")
    bh.alignment = CENTER
    ws.row_dimensions[row].height = 20
    row += 1

    # Column headers for breakdown
    # Layout: Participant | T1P1 | T1P1 pts | T1P2 | T1P2 pts | ... | Total
    breakdown_headers = ["Participant"]
    for tier in TIERS:
        label = tier.replace("Tier ", "T")
        breakdown_headers += [f"{label} Pick 1", "Pts", f"{label} Pick 2", "Pts"]
    breakdown_headers.append("TOTAL")

    tier_col_fills = []
    tf_list = list(tier_fills.values())
    for tf in tf_list:
        tier_col_fills += [tf, tf, tf, tf]  # 4 cols per tier

    for ci, h in enumerate(breakdown_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font  = WHITE_BOLD
        c.fill  = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()
    ws.row_dimensions[row].height = 22
    row += 1

    # Build tier→picks mapping per participant
    for rank, (name, picks, pick_pts, total) in enumerate(scored, 1):
        fill = ALT_FILL if rank % 2 == 0 else WHITE_FILL
        c = ws.cell(row=row, column=1, value=name)
        c.font   = BOLD
        c.fill   = fill
        c.alignment = LEFT
        c.border = _thin_border()

        col = 2
        # picks are ordered: T1P1, T1P2, T2P1, T2P2, ..., T6P1, T6P2
        for ti, tier_fill in enumerate(tf_list):
            for pi in range(2):
                idx = ti * 2 + pi
                team = picks[idx] if idx < len(picks) else ""
                pts  = pick_pts.get(team, 0) if team else ""

                tc = ws.cell(row=row, column=col, value=team)
                tc.fill      = tier_fill
                tc.font      = Font(color="FFFFFF")
                tc.alignment = CENTER
                tc.border    = _thin_border()
                col += 1

                pc = ws.cell(row=row, column=col, value=pts)
                pc.fill      = YELLOW_FILL if pts else fill
                pc.font      = BOLD if pts else Font()
                pc.alignment = CENTER
                pc.border    = _thin_border()
                col += 1

        tc2 = ws.cell(row=row, column=col, value=total)
        tc2.font      = Font(bold=True, size=11)
        tc2.fill      = GREEN_FILL
        tc2.alignment = CENTER
        tc2.border    = _thin_border()
        row += 1

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    for ci in range(2, len(breakdown_headers) + 1):
        h = breakdown_headers[ci - 1] if ci <= len(breakdown_headers) else ""
        if h == "Pts" or h == "TOTAL":
            ws.column_dimensions[get_column_letter(ci)].width = 7
        else:
            ws.column_dimensions[get_column_letter(ci)].width = 16

    # Leaderboard columns
    for ci, w in enumerate([6, 22, 14, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(
            w, int(ws.column_dimensions[get_column_letter(ci)].width or 0)
        )

    ws.freeze_panes = "B4"
    print(f"  Scoring tab: {len(scored)} participant(s) scored, {len(matches)} finished match(es).")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Build Scores and Scoring tabs.")
    parser.add_argument("--scores",  action="store_true", help="Rebuild Scores tab only")
    parser.add_argument("--scoring", action="store_true", help="Rebuild Scoring tab only")
    args = parser.parse_args()

    do_scores  = args.scores  or not args.scoring
    do_scoring = args.scoring or not args.scores

    wb = openpyxl.load_workbook(MASTER_FILE)

    if do_scores:
        if SCORES_SHEET in wb.sheetnames:
            del wb[SCORES_SHEET]
        ws_scores = wb.create_sheet(SCORES_SHEET)
        print(f"Building '{SCORES_SHEET}' tab...")
        build_scores_tab(ws_scores)

    if do_scoring:
        if SCORING_SHEET in wb.sheetnames:
            del wb[SCORING_SHEET]
        ws_scoring = wb.create_sheet(SCORING_SHEET)
        print(f"Building '{SCORING_SHEET}' tab...")
        build_scoring_tab(ws_scoring, wb)

    wb.save(MASTER_FILE)
    print(f"\nSaved → {MASTER_FILE}")


if __name__ == "__main__":
    main()
