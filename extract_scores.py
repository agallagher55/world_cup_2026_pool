"""
Fetches 2026 FIFA World Cup match results and updates WC2026_Pool.xlsx.

Data source: football-data.org free tier (no cost, requires a free API key).
  1. Register at https://www.football-data.org/client/register
  2. Copy your key and either:
       export FOOTBALL_DATA_API_KEY="your_key_here"
     or pass it via --api-key on the command line.

Usage:
  python extract_scores.py                    # live data from API
  python extract_scores.py --mock             # use bundled mock data (no key needed)
  python extract_scores.py --api-key MY_KEY   # pass key directly

What this script does:
  1. Pulls all WC2026 match results from the API (or mock data).
  2. Writes / refreshes a "Match Results" sheet in WC2026_Pool.xlsx.
  3. Reads participant picks from the "Picks" sheet (built by extract_picks.py).
  4. Calculates points per the Scoring Rules sheet and writes totals back to
     the "Picks" sheet's TOTAL PTS column.

Scoring rules applied:
  Group stage:
    Win (regulation)  → 3 pts
    Draw              → 1 pt
    Loss              → 0 pts
  Knockout stage (no draws):
    Win (including AET/pens)  → 3 pts  [modifier: if AET/pens → 2 pts]
    Loss (AET/pens)           → 1 pt
  Group-stage bonuses (applied after group stage is complete):
    Win group                         → +3 pts
    Most goals scored in group        → +3 pts  (+3 more if best in tournament)
    Fewest goals conceded in group    → +3 pts  (+3 more if best in tournament)
"""

import os
import sys
import json
import argparse
from datetime import datetime, timezone

try:
    import requests
except ImportError:
    sys.exit("Install requests:  pip install requests")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Install openpyxl:  pip install openpyxl")

# ── Constants ────────────────────────────────────────────────────────────────

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Works whether the script lives at the project root or in a scripts/ subfolder
_PROJECT_FOLDER = _SCRIPT_DIR if os.path.isfile(os.path.join(_SCRIPT_DIR, "WC2026_Pool.xlsx")) else os.path.dirname(_SCRIPT_DIR)
MASTER_FILE = os.path.join(_PROJECT_FOLDER, "WC2026_Pool.xlsx")
API_BASE = "https://api.football-data.org/v4"
WC_CODE = "WC"  # competition code for FIFA World Cup

TIERS = ["Tier 1", "Tier 2", "Tier 3", "Tier 4", "Tier 5", "Tier 6"]

# Mock data used with --mock flag: a small sample of plausible WC2026 results
# covering each scoring-rule type so the full pipeline can be tested locally.
MOCK_MATCHES = [
    # --- Group A ---
    {
        "id": 1, "stage": "GROUP_STAGE", "group": "GROUP_A", "status": "FINISHED",
        "utcDate": "2026-06-11T20:00:00Z",
        "homeTeam": "Mexico", "awayTeam": "South Korea",
        "homeScore": 2, "awayScore": 1, "winner": "HOME_TEAM",
        "duration": "REGULAR",
    },
    {
        "id": 2, "stage": "GROUP_STAGE", "group": "GROUP_A", "status": "FINISHED",
        "utcDate": "2026-06-11T23:00:00Z",
        "homeTeam": "South Africa", "awayTeam": "Czechia",
        "homeScore": 0, "awayScore": 0, "winner": "DRAW",
        "duration": "REGULAR",
    },
    {
        "id": 3, "stage": "GROUP_STAGE", "group": "GROUP_A", "status": "FINISHED",
        "utcDate": "2026-06-15T20:00:00Z",
        "homeTeam": "Mexico", "awayTeam": "South Africa",
        "homeScore": 3, "awayScore": 0, "winner": "HOME_TEAM",
        "duration": "REGULAR",
    },
    {
        "id": 4, "stage": "GROUP_STAGE", "group": "GROUP_A", "status": "FINISHED",
        "utcDate": "2026-06-15T23:00:00Z",
        "homeTeam": "South Korea", "awayTeam": "Czechia",
        "homeScore": 1, "awayScore": 2, "winner": "AWAY_TEAM",
        "duration": "REGULAR",
    },
    {
        "id": 5, "stage": "GROUP_STAGE", "group": "GROUP_A", "status": "FINISHED",
        "utcDate": "2026-06-19T20:00:00Z",
        "homeTeam": "Mexico", "awayTeam": "Czechia",
        "homeScore": 1, "awayScore": 0, "winner": "HOME_TEAM",
        "duration": "REGULAR",
    },
    {
        "id": 6, "stage": "GROUP_STAGE", "group": "GROUP_A", "status": "FINISHED",
        "utcDate": "2026-06-19T20:00:00Z",
        "homeTeam": "South Africa", "awayTeam": "South Korea",
        "homeScore": 1, "awayScore": 2, "winner": "AWAY_TEAM",
        "duration": "REGULAR",
    },
    # --- Group B (Canada, Switzerland, Qatar, Bosnia) ---
    {
        "id": 7, "stage": "GROUP_STAGE", "group": "GROUP_B", "status": "FINISHED",
        "utcDate": "2026-06-12T17:00:00Z",
        "homeTeam": "Canada", "awayTeam": "Bosnia & Herz.",
        "homeScore": 2, "awayScore": 0, "winner": "HOME_TEAM",
        "duration": "REGULAR",
    },
    {
        "id": 8, "stage": "GROUP_STAGE", "group": "GROUP_B", "status": "FINISHED",
        "utcDate": "2026-06-12T20:00:00Z",
        "homeTeam": "Switzerland", "awayTeam": "Qatar",
        "homeScore": 3, "awayScore": 1, "winner": "HOME_TEAM",
        "duration": "REGULAR",
    },
    {
        "id": 9, "stage": "GROUP_STAGE", "group": "GROUP_B", "status": "FINISHED",
        "utcDate": "2026-06-16T17:00:00Z",
        "homeTeam": "Canada", "awayTeam": "Qatar",
        "homeScore": 4, "awayScore": 0, "winner": "HOME_TEAM",
        "duration": "REGULAR",
    },
    {
        "id": 10, "stage": "GROUP_STAGE", "group": "GROUP_B", "status": "FINISHED",
        "utcDate": "2026-06-16T20:00:00Z",
        "homeTeam": "Switzerland", "awayTeam": "Bosnia & Herz.",
        "homeScore": 2, "awayScore": 2, "winner": "DRAW",
        "duration": "REGULAR",
    },
    {
        "id": 11, "stage": "GROUP_STAGE", "group": "GROUP_B", "status": "FINISHED",
        "utcDate": "2026-06-20T20:00:00Z",
        "homeTeam": "Canada", "awayTeam": "Switzerland",
        "homeScore": 1, "awayScore": 2, "winner": "AWAY_TEAM",
        "duration": "REGULAR",
    },
    {
        "id": 12, "stage": "GROUP_STAGE", "group": "GROUP_B", "status": "FINISHED",
        "utcDate": "2026-06-20T20:00:00Z",
        "homeTeam": "Bosnia & Herz.", "awayTeam": "Qatar",
        "homeScore": 2, "awayScore": 1, "winner": "HOME_TEAM",
        "duration": "REGULAR",
    },
    # --- A knockout match to test AET/pen logic ---
    {
        "id": 100, "stage": "LAST_16", "group": None, "status": "FINISHED",
        "utcDate": "2026-07-01T20:00:00Z",
        "homeTeam": "Spain", "awayTeam": "Canada",
        "homeScore": 1, "awayScore": 1, "winner": "HOME_TEAM",
        "duration": "PENALTY_SHOOTOUT",
    },
    {
        "id": 101, "stage": "LAST_16", "group": None, "status": "FINISHED",
        "utcDate": "2026-07-01T23:00:00Z",
        "homeTeam": "France", "awayTeam": "Switzerland",
        "homeScore": 2, "awayScore": 0, "winner": "HOME_TEAM",
        "duration": "REGULAR",
    },
    # --- A scheduled (future) match to test filtering ---
    {
        "id": 200, "stage": "LAST_16", "group": None, "status": "SCHEDULED",
        "utcDate": "2026-07-03T20:00:00Z",
        "homeTeam": "Argentina", "awayTeam": "Norway",
        "homeScore": None, "awayScore": None, "winner": None,
        "duration": "REGULAR",
    },
]


# ── API helpers ───────────────────────────────────────────────────────────────

def fetch_matches(api_key: str) -> list[dict]:
    """Fetch all WC 2026 matches from football-data.org."""
    headers = {"X-Auth-Token": api_key}
    url = f"{API_BASE}/competitions/{WC_CODE}/matches"
    resp = requests.get(url, headers=headers, timeout=15)
    if resp.status_code == 403:
        sys.exit(
            "API returned 403 Forbidden. Check that your key is valid and that "
            "the free tier covers the WC competition (it does — re-check the key)."
        )
    if resp.status_code == 429:
        sys.exit("Rate limit hit. The free tier allows 10 requests/min — wait a moment and retry.")
    resp.raise_for_status()
    raw_matches = resp.json().get("matches", [])
    return _normalize_api_matches(raw_matches)


def _normalize_api_matches(raw: list[dict]) -> list[dict]:
    """Map football-data.org match structure to our internal format."""
    out = []
    for m in raw:
        score = m.get("score", {})
        full = score.get("fullTime", {})
        duration = score.get("duration", "REGULAR")
        winner = score.get("winner")
        home_score = full.get("home")
        away_score = full.get("away")
        out.append({
            "id": m["id"],
            "stage": m.get("stage", ""),
            "group": m.get("group"),
            "status": m.get("status", ""),
            "utcDate": m.get("utcDate", ""),
            "homeTeam": m["homeTeam"]["name"],
            "awayTeam": m["awayTeam"]["name"],
            "homeScore": home_score,
            "awayScore": away_score,
            "winner": winner,
            "duration": duration,
        })
    return out


# ── Scoring logic ─────────────────────────────────────────────────────────────

GROUP_STAGES = {"GROUP_STAGE"}
KNOCKOUT_STAGES = {"LAST_16", "QUARTER_FINALS", "SEMI_FINALS", "THIRD_PLACE", "FINAL"}


def compute_team_stats(matches: list[dict]) -> dict:
    """
    Returns a dict keyed by team name with:
      points          – match-result points (3/1/0 or 2/1 for knockout)
      group_wins      – wins in group stage
      group_goals_scored
      group_goals_conceded
      group_name      – which group they're in
      matches_played  – finished matches count
    """
    stats: dict[str, dict] = {}

    def get(team):
        if team not in stats:
            stats[team] = {
                "points": 0,
                "group_wins": 0,
                "group_goals_scored": 0,
                "group_goals_conceded": 0,
                "group_name": None,
                "matches_played": 0,
            }
        return stats[team]

    for m in matches:
        if m["status"] != "FINISHED":
            continue
        home, away = m["homeTeam"], m["awayTeam"]
        hs, as_ = m["homeScore"], m["awayScore"]
        winner = m["winner"]
        stage = m["stage"]
        duration = m["duration"]
        group = m.get("group")

        if hs is None or as_ is None:
            continue

        home_s = get(home)
        away_s = get(away)
        home_s["matches_played"] += 1
        away_s["matches_played"] += 1

        if group:
            home_s["group_name"] = group
            away_s["group_name"] = group

        if stage in GROUP_STAGES:
            home_s["group_goals_scored"] += hs
            home_s["group_goals_conceded"] += as_
            away_s["group_goals_scored"] += as_
            away_s["group_goals_conceded"] += hs

            if winner == "HOME_TEAM":
                home_s["points"] += 3
                home_s["group_wins"] += 1
            elif winner == "AWAY_TEAM":
                away_s["points"] += 3
                away_s["group_wins"] += 1
            else:  # DRAW
                home_s["points"] += 1
                away_s["points"] += 1

        elif stage in KNOCKOUT_STAGES:
            aet = duration in ("EXTRA_TIME", "PENALTY_SHOOTOUT")
            if winner == "HOME_TEAM":
                home_s["points"] += 2 if aet else 3
                away_s["points"] += 1 if aet else 0
            elif winner == "AWAY_TEAM":
                away_s["points"] += 2 if aet else 3
                home_s["points"] += 1 if aet else 0

    return stats


def apply_group_bonuses(stats: dict) -> dict:
    """
    Adds group-stage bonus points to stats in-place.
    Returns a copy annotated with bonus breakdown for reporting.
    """
    # Collect teams per group
    groups: dict[str, list[str]] = {}
    for team, s in stats.items():
        g = s["group_name"]
        if g:
            groups.setdefault(g, []).append(team)

    all_scored = {t: stats[t]["group_goals_scored"] for t in stats if stats[t]["group_name"]}
    all_conceded = {t: stats[t]["group_goals_conceded"] for t in stats if stats[t]["group_name"]}

    if not all_scored:
        return stats

    tournament_best_scored = max(all_scored.values())
    tournament_best_conceded = min(all_conceded.values())

    for group_name, teams in groups.items():
        # Group winner(s) — max wins, then points as tiebreaker
        max_wins = max(stats[t]["group_wins"] for t in teams)
        group_winners = [t for t in teams if stats[t]["group_wins"] == max_wins]

        # Most goals scored in group
        max_group_scored = max(stats[t]["group_goals_scored"] for t in teams)
        top_scorers = [t for t in teams if stats[t]["group_goals_scored"] == max_group_scored]

        # Fewest goals conceded in group
        min_group_conceded = min(stats[t]["group_goals_conceded"] for t in teams)
        best_defense = [t for t in teams if stats[t]["group_goals_conceded"] == min_group_conceded]

        for team in teams:
            s = stats[team]
            s.setdefault("bonuses", {})
            if team in group_winners:
                s["points"] += 3
                s["bonuses"]["group_winner"] = 3
            if team in top_scorers:
                s["points"] += 3
                s["bonuses"]["top_scorer_group"] = 3
                if stats[team]["group_goals_scored"] >= tournament_best_scored:
                    s["points"] += 3
                    s["bonuses"]["top_scorer_tournament"] = 3
            if team in best_defense:
                s["points"] += 3
                s["bonuses"]["best_defense_group"] = 3
                if stats[team]["group_goals_conceded"] <= tournament_best_conceded:
                    s["points"] += 3
                    s["bonuses"]["best_defense_tournament"] = 3

    return stats


def score_participants(picks_by_participant: dict, team_stats: dict) -> dict:
    """Returns {participant: total_points}."""
    totals = {}
    for participant, picks in picks_by_participant.items():
        total = 0
        for team in picks:
            if team and team in team_stats:
                total += team_stats[team].get("points", 0)
        totals[participant] = total
    return totals


# ── Spreadsheet I/O ────────────────────────────────────────────────────────────

def read_participant_picks() -> dict[str, list[str]]:
    """Read the Picks sheet and return {participant: [team, ...]}."""
    wb = openpyxl.load_workbook(MASTER_FILE)
    if "Picks" not in wb.sheetnames:
        sys.exit("No 'Picks' sheet found — run extract_picks.py first.")
    ws = wb["Picks"]

    participants = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[0]
        if not name:
            continue
        # columns 2–13 are the 12 picks (Tier 1 P1, T1 P2, T2 P1 … T6 P2)
        picks = [str(row[i]).strip() for i in range(1, 13) if row[i]]
        participants[name] = picks
    return participants


def write_match_results_sheet(wb: openpyxl.Workbook, matches: list[dict]):
    """Write / refresh the Match Results sheet."""
    if "Match Results" in wb.sheetnames:
        del wb["Match Results"]
    ws = wb.create_sheet("Match Results")

    header_fill = PatternFill("solid", fgColor="1F4E79")
    white_bold = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Date (UTC)", "Stage", "Group", "Home Team", "Score", "Away Team", "Winner", "Duration", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    finished = [m for m in matches if m["status"] == "FINISHED"]
    scheduled = [m for m in matches if m["status"] != "FINISHED"]
    ordered = sorted(finished, key=lambda m: m["utcDate"]) + sorted(scheduled, key=lambda m: m["utcDate"])

    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    for row_idx, m in enumerate(ordered, start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        date_str = m["utcDate"][:10] if m["utcDate"] else ""
        score_str = (
            f"{m['homeScore']} – {m['awayScore']}"
            if m["homeScore"] is not None and m["awayScore"] is not None
            else "vs"
        )
        row_vals = [
            date_str,
            m["stage"],
            m["group"] or "",
            m["homeTeam"],
            score_str,
            m["awayTeam"],
            m["winner"] or "",
            m["duration"],
            m["status"],
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = center
            cell.fill = fill
            cell.border = border

    col_widths = [12, 16, 10, 20, 10, 20, 14, 20, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"


def write_scores_to_scoring_sheet(wb: openpyxl.Workbook, totals: dict[str, int]):
    """Update TOTAL PTS column in the Picks sheet."""
    ws = wb["Picks"]
    # Find TOTAL PTS column (last data column)
    total_col = None
    for cell in ws[2]:
        if cell.value and "TOTAL" in str(cell.value):
            total_col = cell.column
            break
    if not total_col:
        print("  WARNING: could not find TOTAL PTS column in Picks sheet.")
        return

    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        name = row[0].value
        if name and name in totals:
            cell = ws.cell(row=row[0].row, column=total_col, value=totals[name])
            cell.fill = yellow_fill
            cell.font = bold
            cell.alignment = center


def write_team_stats_sheet(wb: openpyxl.Workbook, team_stats: dict):
    """Write / refresh a Team Stats sheet for debugging and transparency."""
    if "Team Stats" in wb.sheetnames:
        del wb["Team Stats"]
    ws = wb.create_sheet("Team Stats")

    header_fill = PatternFill("solid", fgColor="1F4E79")
    white_bold = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Team", "Group", "Matches Played",
        "Gp Goals Scored", "Gp Goals Conceded",
        "Match Pts", "Bonus Pts", "Total Pts",
        "Bonus Breakdown",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    sorted_teams = sorted(team_stats.items(), key=lambda x: -x[1]["points"])
    for row_idx, (team, s) in enumerate(sorted_teams, start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        bonuses = s.get("bonuses", {})
        bonus_total = sum(bonuses.values())
        match_pts = s["points"] - bonus_total
        row_vals = [
            team,
            s["group_name"] or "",
            s["matches_played"],
            s["group_goals_scored"],
            s["group_goals_conceded"],
            match_pts,
            bonus_total,
            s["points"],
            ", ".join(f"{k}(+{v})" for k, v in bonuses.items()) if bonuses else "",
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = center
            cell.fill = fill
            cell.border = border

    col_widths = [22, 10, 16, 18, 20, 10, 10, 10, 45]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Fetch WC2026 scores and update pool spreadsheet.")
    parser.add_argument("--mock", action="store_true", help="Use built-in mock data instead of live API.")
    parser.add_argument("--api-key", default=os.environ.get("FOOTBALL_DATA_API_KEY", ""), help="football-data.org API key.")
    parser.add_argument("--no-bonuses", action="store_true", help="Skip group-stage bonus calculation.")
    args = parser.parse_args()

    if args.mock:
        print("Using mock match data (--mock flag set).")
        matches = MOCK_MATCHES
    else:
        if not args.api_key:
            sys.exit(
                "No API key provided.\n"
                "  Set FOOTBALL_DATA_API_KEY env var, or pass --api-key YOUR_KEY.\n"
                "  Get a free key at https://www.football-data.org/client/register\n"
                "  Or test without a key using --mock."
            )
        print("Fetching live match data from football-data.org...")
        matches = fetch_matches(args.api_key)

    finished = [m for m in matches if m["status"] == "FINISHED"]
    scheduled = [m for m in matches if m["status"] != "FINISHED"]
    print(f"  {len(matches)} total matches | {len(finished)} finished | {len(scheduled)} pending/scheduled")

    print("Computing team stats...")
    team_stats = compute_team_stats(matches)

    if not args.no_bonuses:
        print("Applying group-stage bonuses...")
        team_stats = apply_group_bonuses(team_stats)

    print("Reading participant picks from Picks sheet...")
    picks = read_participant_picks()
    if not picks:
        sys.exit("No participant picks found — run extract_picks.py first.")

    print("Calculating participant totals...")
    totals = score_participants(picks, team_stats)

    print("\nStandings so far:")
    for participant, pts in sorted(totals.items(), key=lambda x: -x[1]):
        picked_teams = picks[participant]
        team_breakdown = []
        for team in picked_teams:
            pts_for_team = team_stats.get(team, {}).get("points", 0)
            team_breakdown.append(f"{team}({pts_for_team})")
        print(f"  {participant:20s}  {pts:3d} pts  [{', '.join(team_breakdown)}]")

    print(f"\nUpdating {MASTER_FILE}...")
    wb = openpyxl.load_workbook(MASTER_FILE)
    write_match_results_sheet(wb, matches)
    write_scores_to_scoring_sheet(wb, totals)
    write_team_stats_sheet(wb, team_stats)
    wb.save(MASTER_FILE)
    print("Done. Sheets updated: Match Results, Picks (TOTAL PTS), Team Stats.")


if __name__ == "__main__":
    main()
